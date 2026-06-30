"""Lector multi-formato de adjuntos de pliegos: PDF, Excel (.xlsx) y Word (.docx).

Devuelve texto plano y la lista de renglones detectados, reutilizando el
detector de renglones de `parsing.pdf`. Cada formato se maneja con dependencias
livianas y con import diferido, así el paquete importa aunque falte alguna.

Notas:
  - .xlsx: openpyxl (ya está en requirements).
  - .docx: se parsea el XML interno con zipfile (sin dependencia extra).
  - .pdf: parsing.pdf.
  - .xls / .doc (binario viejo OLE): best-effort; si no hay librería, se saltea
    con aviso (la mayoría de los pliegos de CODINEU son PDF).
"""

from __future__ import annotations

import re
import zipfile
from pathlib import Path

from radar.logging_setup import get_logger
from radar.parsing.pdf import detectar_renglones, leer_pdf

logger = get_logger("parsing.documents")


def _limpiar(t: str) -> str:
    return re.sub(r"\s+", " ", (t or "").replace("\n", " ")).strip()


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def leer_xlsx(path: str | Path) -> str:
    try:
        from openpyxl import load_workbook  # import diferido
    except ImportError:
        logger.warning("openpyxl no instalado; no se puede leer %s", path)
        return ""
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        filas: list[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                celdas = [str(c) for c in row if c is not None and str(c).strip()]
                if celdas:
                    filas.append(" ; ".join(celdas))
        wb.close()
        return "\n".join(filas)
    except Exception as e:  # noqa: BLE001
        logger.error("Error leyendo XLSX %s: %s", path, e)
        return ""


# --------------------------------------------------------------------------- #
# Word
# --------------------------------------------------------------------------- #
def leer_docx(path: str | Path) -> str:
    """Extrae texto de un .docx leyendo word/document.xml (sin python-docx)."""
    try:
        with zipfile.ZipFile(str(path)) as z:
            if "word/document.xml" not in z.namelist():
                return ""
            xml = z.read("word/document.xml").decode("utf-8", errors="ignore")
    except Exception as e:  # noqa: BLE001
        logger.error("Error abriendo DOCX %s: %s", path, e)
        return ""

    # Convertir saltos de párrafo/fila a separadores antes de quitar tags.
    xml = re.sub(r"</w:p>", "\n", xml)
    xml = re.sub(r"</w:tr>", "\n", xml)
    xml = re.sub(r"<w:tab[^>]*/>", " ; ", xml)
    textos = re.findall(r"<w:t[^>]*>(.*?)</w:t>", xml, flags=re.S)
    # Re-armar respetando líneas: reconstruimos desde el xml ya con \n.
    sin_tags = re.sub(r"<[^>]+>", "", xml)
    if textos and not sin_tags.strip():
        sin_tags = " ".join(textos)
    lineas = [ _limpiar(l) for l in sin_tags.splitlines() ]
    return "\n".join(l for l in lineas if l)


# --------------------------------------------------------------------------- #
# Router
# --------------------------------------------------------------------------- #
_EXT_NO_SOPORTADAS = {".xls", ".doc", ".rar"}


def leer_documento(path: str | Path) -> dict:
    """Lee un adjunto y devuelve {tipo, texto, renglones}."""
    path = Path(path)
    ext = path.suffix.lower()

    if ext == ".pdf":
        texto = leer_pdf(path)
        tipo = "pdf"
    elif ext == ".xlsx":
        texto = leer_xlsx(path)
        tipo = "xlsx"
    elif ext == ".docx":
        texto = leer_docx(path)
        tipo = "docx"
    elif ext in _EXT_NO_SOPORTADAS:
        logger.info("Formato %s no soportado sin libs extra: %s", ext, path.name)
        return {"tipo": ext.lstrip("."), "texto": "", "renglones": []}
    else:
        # Intento como texto plano.
        try:
            texto = path.read_text(encoding="utf-8", errors="ignore")
            tipo = "txt"
        except Exception:  # noqa: BLE001
            return {"tipo": ext.lstrip(".") or "?", "texto": "", "renglones": []}

    renglones = detectar_renglones(texto) if texto else []
    return {"tipo": tipo, "texto": texto, "renglones": renglones}


def leer_documentos(paths: list[str | Path]) -> list[dict]:
    salida = []
    for p in paths or []:
        info = leer_documento(p)
        info["archivo"] = str(p)
        salida.append(info)
    return salida
